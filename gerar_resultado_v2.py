from openpyxl import load_workbook, Workbook
from openpyxl.styles import Border, Side, PatternFill
from statistics import mean, stdev, median
from datetime import datetime
from os import listdir

thin_border = Border(
    left = Side(style = 'thin'), 
    right = Side(style = 'thin'), 
    top = Side(style = 'thin'), 
    bottom = Side(style = 'thin')
)

graybg = PatternFill(start_color = 'cccccc', fill_type = "solid")

def color_row(row_tuple, color):
    background = PatternFill(start_color = color, fill_type = "solid")
    for cell in row_tuple:
        cell.fill = background
        cell.border = thin_border


def copy_sheet(ws_source, ws_destination):
    mr = ws_source.max_row
    mc = ws_source.max_column
    
    for i in range (1, mr + 1):
        for j in range (1, mc + 1):
            c = ws_source.cell(row = i, column = j)
            ws_destination.cell(row = i, column = j).value = c.value



resultado = Workbook()
itens = listdir('itens')

deve_conter = {'item1':['CANETA', 'MARCA-TEXTO'], 'item2':['CANETA', 'MARCA-TEXTO']}
proibidas = {'item1':['P3', 'P4'], 'item2':['P1', 'P2']}
unid_forn = {'item1':['UNIDADE'], 'item2':['UNIDADE', 'CAIXA 12,00 UN']}
cod_mat = {'item1':[279313], 'item2':[279313]}
periodo = {'item1':['07/2020', '07/2021'], 'item2':['10/2020', '05/2021']}

for item in itens:
    wb = load_workbook(f'itens/{item}')
    ws = wb.active

    lastrow = ws.max_row
    item_name = item[0:-5]

    deve_conter_item = deve_conter[item_name]
    proibidas_item = proibidas[item_name]
    unid_forn_item = unid_forn[item_name]
    cod_mat_item = cod_mat[item_name]
    periodo_item = periodo[item_name]

    data_inicial = datetime.strptime(periodo_item[0], '%m/%Y')
    data_final = datetime.strptime(periodo_item[1], '%m/%Y')
    
    ws['M5'] = 'Item Ativo'
    ws['M5'].fill = graybg

    for row, i in zip(ws.iter_rows(min_row = 6, max_row = lastrow, max_col = 12), range(6, lastrow + 1)):
        descr_cell = row[4]
        cod_mat_cell = row[3]
        unid_forn_cell = row[5]
        data_cell = row[11]
        
        descr = descr_cell.value
        cod = cod_mat_cell.value
        unid = unid_forn_cell.value
        data = datetime.strptime(data_cell.value, '%d/%m/%Y')
        if any(x in descr for x in proibidas_item):
            ws[f'M{i}'] = 0
        elif all(x in descr for x in deve_conter_item) and (data_inicial <= data <= data_final) and (cod in cod_mat_item) and (unid in unid_forn_item):
            ws[f'M{i}'] = 1
        else:
            ws[f'M{i}'] = 0

    
    unit_values = [float(cell[0].value.replace(',', '.')) for cell in ws['H6':f'H{lastrow}']]

    valores = []
    for row, i in zip(ws.iter_rows(min_row = 6, max_row = lastrow, min_col = 13, max_col = 13), range(0, lastrow - 5)):
        cell = row[0]
        ativo = cell.value
        if ativo == 1:
            valores.append(unit_values[i])
        
    media = mean(valores)
    desvio = stdev(valores)
    coeficiente = desvio / media
    mediana = median(valores)
    preco = mediana if coeficiente > 0.25 else media

    ws[f'A{lastrow + 2}'] = 'Média'
    ws[f'A{lastrow + 3}'] = 'Desvio'
    ws[f'A{lastrow + 4}'] = 'Coeficiente'
    ws[f'A{lastrow + 5}'] = 'Mediana'
    ws[f'A{lastrow + 6}'] = 'Preço'

    ws[f'B{lastrow + 2}'] = media
    ws[f'B{lastrow + 3}'] = desvio
    ws[f'B{lastrow + 4}'] = coeficiente
    ws[f'B{lastrow + 5}'] = mediana
    ws[f'B{lastrow + 6}'] = preco

    ws_result = resultado.create_sheet(item_name)
    copy_sheet(ws, ws_result)

    for row in ws_result.iter_rows(min_row = 6, max_row = lastrow, max_col = 13):
        cell = row[12]
        ativo = cell.value
        if ativo == 1:
            color_row(row, 'e2f0d9')
        else:
            color_row(row, 'fbe5d6')

    ws_result['A5'].fill = graybg
    ws_result['B5'].fill = graybg
    ws_result['C5'].fill = graybg
    ws_result['D5'].fill = graybg
    ws_result['E5'].fill = graybg
    ws_result['F5'].fill = graybg
    ws_result['G5'].fill = graybg
    ws_result['H5'].fill = graybg
    ws_result['I5'].fill = graybg
    ws_result['J5'].fill = graybg
    ws_result['K5'].fill = graybg
    ws_result['L5'].fill = graybg
    ws_result['M5'].fill = graybg

    ws_result['A5'].border = thin_border
    ws_result['B5'].border = thin_border
    ws_result['C5'].border = thin_border
    ws_result['D5'].border = thin_border
    ws_result['E5'].border = thin_border
    ws_result['F5'].border = thin_border
    ws_result['G5'].border = thin_border
    ws_result['H5'].border = thin_border
    ws_result['I5'].border = thin_border
    ws_result['J5'].border = thin_border
    ws_result['K5'].border = thin_border
    ws_result['L5'].border = thin_border
    ws_result['M5'].border = thin_border

    ws_result[f'A{lastrow + 2}'].border = thin_border
    ws_result[f'A{lastrow + 3}'].border = thin_border
    ws_result[f'A{lastrow + 4}'].border = thin_border
    ws_result[f'A{lastrow + 5}'].border = thin_border
    ws_result[f'A{lastrow + 6}'].border = thin_border
    ws_result[f'B{lastrow + 2}'].border = thin_border
    ws_result[f'B{lastrow + 3}'].border = thin_border
    ws_result[f'B{lastrow + 4}'].border = thin_border
    ws_result[f'B{lastrow + 5}'].border = thin_border
    ws_result[f'B{lastrow + 6}'].border = thin_border


rm_sheet = resultado['Sheet']
resultado.remove(rm_sheet)
resultado.save('resultado.xlsx')