from openpyxl import load_workbook
from openpyxl.styles import Border, Side, PatternFill
from statistics import mean, stdev, median
from datetime import date

thin_border = Border(
    left = Side(style = 'thin'), 
    right = Side(style = 'thin'), 
    top = Side(style = 'thin'), 
    bottom = Side(style = 'thin')
)

def color_row(row_tuple, color):
    background = PatternFill(start_color = color, fill_type = "solid")
    for cell in row_tuple:
        cell.fill = background

wb = load_workbook('relatorio_painel.xlsx')
ws = wb.active

lastrow = ws.max_row

deve_conter = ['CANETA', 'MARCA-TEXTO']
proibidas = ['P3', 'P4']

for row in ws.iter_rows(min_row = 6, max_row = lastrow, max_col = 12):
    descr = row[4].value
    if any(x in descr for x in proibidas):
        continue

    if all(x in descr for x in deve_conter):
        color_row(row, 'e2f0d9')

unit_values = [float(cell[0].value.replace(',', '.')) for cell in ws['H6':f'H{lastrow}']]

mes_atual = date.today().month
ano_atual = date.today().year

valores = []
for row, i in zip(ws.iter_rows(min_row = 6, max_row = lastrow, min_col = 12, max_col = 12), range(0, lastrow-5)):
    cell = row[0]
    data = cell.value
    bgcolor = cell.fill.start_color.index
    mes = int(data.split('/')[1])
    ano = int(data.split('/')[2])
    if bgcolor == '00e2f0d9':
        if ano == 2020 and mes in range(mes_atual, 13):
            valores.append(unit_values[i])

        elif ano == 2021 and mes in range(mes_atual + 1):
            valores.append(unit_values[i])
        
media = mean(valores)
desvio = stdev(valores)
coeficiente = desvio / media
mediana = median(valores)
preco = mediana if coeficiente > 0.25 else media

ws[f'A{lastrow + 2}'] = 'Média'
ws[f'A{lastrow + 3}'] = 'Desvio'
ws[f'A{lastrow + 4}'] = 'Coeficiente'
ws[f'A{lastrow + 5}'] = 'Mediana'
ws[f'A{lastrow + 6}'] = 'Preço'

ws[f'B{lastrow + 2}'] = media
ws[f'B{lastrow + 3}'] = desvio
ws[f'B{lastrow + 4}'] = coeficiente
ws[f'B{lastrow + 5}'] = mediana
ws[f'B{lastrow + 6}'] = preco

ws[f'A{lastrow + 2}'].border = thin_border
ws[f'A{lastrow + 3}'].border = thin_border
ws[f'A{lastrow + 4}'].border = thin_border
ws[f'A{lastrow + 5}'].border = thin_border
ws[f'A{lastrow + 6}'].border = thin_border
ws[f'B{lastrow + 2}'].border = thin_border
ws[f'B{lastrow + 3}'].border = thin_border
ws[f'B{lastrow + 4}'].border = thin_border
ws[f'B{lastrow + 5}'].border = thin_border
ws[f'B{lastrow + 6}'].border = thin_border

wb.save('resultado.xlsx')
