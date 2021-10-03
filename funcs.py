from openpyxl.styles import Border, Side, PatternFill
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter
from unicodedata import normalize

thin_border = Border(
    left = Side(style = 'thin'), 
    right = Side(style = 'thin'), 
    top = Side(style = 'thin'), 
    bottom = Side(style = 'thin')
)

gray_bg = PatternFill(start_color = 'cccccc', fill_type = 'solid')

# Remove a acentuação de uma string
def remove_acc(word: str):
    norm_word = normalize('NFKD', word).encode('ASCII', 'ignore').decode('ASCII')
    return norm_word

# Cria o array de strings (filtros) a partir de uma string
def create_filter(value: str, num: bool = False):
    temp_filters = remove_acc(value.upper()).split('|')
    filters = map(lambda x: x.strip(), temp_filters)

    if num:
        return list(map(lambda x: int(x), filters))

    return list(filters)

# Aplica os filtros em uma descrição
def apply_filter(descr: str, words: list[str], func):
    if words:
        return func(word in descr for word in words)
    else:
        return True

# Colore toda uma linha da tabela
def color_row(row: tuple, color: str):
    background = PatternFill(start_color = color, fill_type = 'solid')
    for cell in row:
        cell.fill = background
        cell.border = thin_border

# Copia todos os elementos de uma work sheet para outra
def copy_sheet(ws_source: Worksheet, ws_destination: Worksheet):
    mr = ws_source.max_row
    mc = ws_source.max_column
    for i in range(1, mr + 1):
        for j in range(1, mc + 1):
            c = ws_source.cell(row = i, column = j)
            ws_destination.cell(row = i, column = j).value = c.value

# Ordena toda a tabela a partir dos valores de uma coluna
def sort_col(ws: Worksheet, col: int, min_row: int, max_row:int, reverse: bool = True):
    col_values = []
    for row in range(min_row, max_row + 1):
        col_values.append(ws.cell(row = row, column = col).value)

    ordered_idx = sorted(range(min_row, max_row + 1), key = lambda k: col_values[k - min_row], reverse = reverse)

    last_col = get_column_letter(ws.max_column)
    ws.move_range(f"A{min_row}:{last_col}{max_row}", max_row)

    for src, idx in zip(ordered_idx, range(min_row, max_row + 1)):
        dest = idx - (src + max_row)
        ws.move_range(f"A{src + max_row}:{last_col}{src + max_row}", dest)