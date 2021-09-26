from openpyxl import load_workbook, Workbook
from statistics import mean, pstdev, median
from datetime import datetime
from os import listdir
import pandas as pd
from funcs import *

resultado = Workbook()
itens = listdir('itens')
filtros = pd.read_excel('filtros.xlsx', converters = {'CÓDIGO DO MATERIAL': str})

flt_item = filtros['ITEM (NOME DO ARQUIVO)']
flt_descr_padrao = filtros['DESCRIÇÃO PADRÃO']
flt_obrigatorias = filtros['DESCRIÇÃO: PALAVRA(S) OBRIGATÓRIA(S)']
flt_deve_conter = filtros['DESCRIÇÃO: DEVE CONTER (MIN 1)']
flt_proibidas = filtros['DESCRIÇÃO: PALAVRA(S) PROIBIDAS(S)']
flt_unid_forn = filtros['UNIDADE DE FORNECIMENTO']
flt_cod_mat = filtros['CÓDIGO DO MATERIAL']
flt_periodo = filtros['PERÍODO']

descr_padrao = {key: value.upper() for key, value in zip(flt_item, flt_descr_padrao)}
obrigatorias = {key: create_filter(value) if isinstance(value, str) else [] for key, value in zip(flt_item, flt_obrigatorias)}
deve_conter = {key: create_filter(value) if isinstance(value, str) else [] for key, value in zip(flt_item, flt_deve_conter)}
proibidas = {key: create_filter(value) if isinstance(value, str) else [] for key, value in zip(flt_item, flt_proibidas)}
unid_forn = {key: create_filter(value) if isinstance(value, str) else [] for key, value in zip(flt_item, flt_unid_forn)}
cod_mat = {key: create_filter(value, num = True) if isinstance(value, str) else [] for key, value in zip(flt_item, flt_cod_mat)}
periodo = {key: create_filter(value) if isinstance(value, str) else [] for key, value in zip(flt_item, flt_periodo)}

for item in itens:
    wb = load_workbook(f'itens/{item}')
    ws = wb.active

    lastrow = ws.max_row
    item_name = item[0:-5]

    descr_padrao_item = descr_padrao[item_name]
    obrigatorias_item = obrigatorias[item_name]
    deve_conter_item = deve_conter[item_name]
    proibidas_item = proibidas[item_name]
    unid_forn_item = unid_forn[item_name]
    cod_mat_item = cod_mat[item_name]
    periodo_item = periodo[item_name]

    ws['C1'] = descr_padrao_item

    if periodo_item:
        data_inicial = datetime.strptime(periodo_item[0], '%m/%Y')
        data_final = datetime.strptime(periodo_item[1], '%m/%Y')
    else:
        data_inicial = datetime.min
        data_final = datetime.max
    
    ws['M5'] = 'Item Ativo'
    ws['M5'].fill = graybg

    for row, i in zip(ws.iter_rows(min_row = 6, max_row = lastrow, max_col = 12), range(6, lastrow + 1)):
        descr_cell = row[4]
        cod_mat_cell = row[3]
        unid_forn_cell = row[5]
        data_cell = row[11]
        
        descr = remove_acc(descr_cell.value).upper().strip()
        cod = cod_mat_cell.value
        unid = remove_acc(unid_forn_cell.value).upper().strip()
        data = datetime.strptime(data_cell.value, '%d/%m/%Y')

        if any(x in descr for x in proibidas_item):
            ws[f'M{i}'] = 0
        elif cod_mat_item:
            if apply_filter(descr, obrigatorias_item, all) \
                and apply_filter(descr, deve_conter_item, any) \
                and apply_filter(unid, unid_forn_item, any) \
                and (data_inicial <= data <= data_final) \
                and (cod in cod_mat_item):

                ws[f'M{i}'] = 1
            else:
                ws[f'M{i}'] = 0
        else:
            if apply_filter(descr, obrigatorias_item, all) \
                and apply_filter(descr, deve_conter_item, any) \
                and apply_filter(unid, unid_forn_item, any) \
                and (data_inicial <= data <= data_final):

                ws[f'M{i}'] = 1
            else:
                ws[f'M{i}'] = 0
            
    
    unit_values = [float(cell[0].value.replace(',', '.')) for cell in ws['H6':f'H{lastrow}']]

    valores = []
    for row, i in zip(ws.iter_rows(min_row = 6, max_row = lastrow, min_col = 13, max_col = 13), range(0, lastrow - 5)):
        cell = row[0]
        ativo = cell.value
        if ativo == 1:
            valores.append(unit_values[i])
        
    media = mean(valores)
    desvio = pstdev(valores)
    coeficiente = desvio / media
    mediana = median(valores)
    preco = mediana if coeficiente > 0.25 else media
    br_supply = preco * 1.11

    ws[f'A{lastrow + 2}'] = 'Média'
    ws[f'A{lastrow + 3}'] = 'Desvio'
    ws[f'A{lastrow + 4}'] = 'Coeficiente'
    ws[f'A{lastrow + 5}'] = 'Mediana'
    ws[f'A{lastrow + 6}'] = 'Preço'
    ws[f'A{lastrow + 7}'] = 'Preço BR Supply'

    ws[f'B{lastrow + 2}'] = media
    ws[f'B{lastrow + 3}'] = desvio
    ws[f'B{lastrow + 4}'] = coeficiente
    ws[f'B{lastrow + 5}'] = mediana
    ws[f'B{lastrow + 6}'] = preco
    ws[f'B{lastrow + 7}'] = br_supply

    ws_result = resultado.create_sheet(item_name)
    copy_sheet(ws, ws_result)

    for row, i in zip(ws_result.iter_rows(min_row = 6, max_row = lastrow, max_col = 13), range(0, lastrow - 5)):
        cell_ativo = row[12]
        cell_valor = row[7]
        cell_valor.value = unit_values[i]
        ativo = cell_ativo.value
        if ativo == 1:
            color_row(row, 'e2f0d9')
        else:
            color_row(row, 'fbe5d6')

    ws_result['A5'].fill = graybg
    ws_result['B5'].fill = graybg
    ws_result['C5'].fill = graybg
    ws_result['D5'].fill = graybg
    ws_result['E5'].fill = graybg
    ws_result['F5'].fill = graybg
    ws_result['G5'].fill = graybg
    ws_result['H5'].fill = graybg
    ws_result['I5'].fill = graybg
    ws_result['J5'].fill = graybg
    ws_result['K5'].fill = graybg
    ws_result['L5'].fill = graybg
    ws_result['M5'].fill = graybg

    ws_result['A5'].border = thin_border
    ws_result['B5'].border = thin_border
    ws_result['C5'].border = thin_border
    ws_result['D5'].border = thin_border
    ws_result['E5'].border = thin_border
    ws_result['F5'].border = thin_border
    ws_result['G5'].border = thin_border
    ws_result['H5'].border = thin_border
    ws_result['I5'].border = thin_border
    ws_result['J5'].border = thin_border
    ws_result['K5'].border = thin_border
    ws_result['L5'].border = thin_border
    ws_result['M5'].border = thin_border

    ws_result[f'A{lastrow + 2}'].border = thin_border
    ws_result[f'A{lastrow + 3}'].border = thin_border
    ws_result[f'A{lastrow + 4}'].border = thin_border
    ws_result[f'A{lastrow + 5}'].border = thin_border
    ws_result[f'A{lastrow + 6}'].border = thin_border
    ws_result[f'A{lastrow + 7}'].border = thin_border
    ws_result[f'B{lastrow + 2}'].border = thin_border
    ws_result[f'B{lastrow + 3}'].border = thin_border
    ws_result[f'B{lastrow + 4}'].border = thin_border
    ws_result[f'B{lastrow + 5}'].border = thin_border
    ws_result[f'B{lastrow + 6}'].border = thin_border
    ws_result[f'B{lastrow + 7}'].border = thin_border


rm_sheet = resultado['Sheet']
resultado.remove(rm_sheet)
resultado.save('resultado.xlsx')